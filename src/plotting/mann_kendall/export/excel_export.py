"""
Excel Export Functions

This module contains functions for exporting analysis results to Excel files.
"""

import pandas as pd
import os
import traceback
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils.dataframe import dataframe_to_rows


def create_analysis_summary_excel(data, well_columns, mapping_dict=None, save_directory='analysis_summary'):
    """
    Create a comprehensive Excel summary file showing which wells were processed/skipped for each analysis type.
    
    Args:
        data: DataFrame with groundwater data
        well_columns: List of well column names
        mapping_dict: Dictionary mapping well names to display names
        save_directory: Directory to save the summary file
    
    Returns:
        str: Path to the created Excel file
    """
    try:
        # Create save directory
        os.makedirs(save_directory, exist_ok=True)
        
        # Initialize summary data
        summary_data = []
        
        for well in well_columns:
            if well in data.columns:
                well_data = data[well].dropna()
                well_display = mapping_dict.get(well, well) if mapping_dict else well
                
                # Basic data metrics
                data_points = len(well_data)
                data_years = 0
                start_date = None
                end_date = None
                
                if data_points > 0 and isinstance(well_data.index, pd.DatetimeIndex):
                    data_years = (well_data.index.max() - well_data.index.min()).days / 365.25
                    start_date = well_data.index.min()
                    end_date = well_data.index.max()
                
                summary_data.append({
                    'Well': well_display,
                    'Original_Name': well,
                    'Total_Data_Points': data_points,
                    'Data_Years': round(data_years, 2) if data_years > 0 else 0,
                    'Start_Date': start_date.strftime('%Y-%m-%d') if start_date else 'N/A',
                    'End_Date': end_date.strftime('%Y-%m-%d') if end_date else 'N/A',
                    'Mean_Level': round(well_data.mean(), 3) if data_points > 0 else 'N/A',
                    'Min_Level': round(well_data.min(), 3) if data_points > 0 else 'N/A',
                    'Max_Level': round(well_data.max(), 3) if data_points > 0 else 'N/A',
                    'Std_Level': round(well_data.std(), 3) if data_points > 0 else 'N/A'
                })
            else:
                well_display = mapping_dict.get(well, well) if mapping_dict else well
                summary_data.append({
                    'Well': well_display,
                    'Original_Name': well,
                    'Total_Data_Points': 0,
                    'Data_Years': 0,
                    'Start_Date': 'N/A',
                    'End_Date': 'N/A',
                    'Mean_Level': 'N/A',
                    'Min_Level': 'N/A',
                    'Max_Level': 'N/A',
                    'Std_Level': 'N/A'
                })
        
        # Create Excel workbook
        wb = Workbook()
        
        # Remove default sheet
        wb.remove(wb.active)
        
        # 1. OVERVIEW SHEET
        ws_overview = wb.create_sheet("Overview")
        overview_data = [
            ['Analysis Summary Report', ''],
            ['Generated on', pd.Timestamp.now().strftime('%Y-%m-%d %H:%M:%S')],
            ['Total Wells in Dataset', len(well_columns)],
            ['Wells with Data', len([w for w in well_columns if w in data.columns and len(data[w].dropna()) > 0])],
            ['Wells without Data', len([w for w in well_columns if w not in data.columns or len(data[w].dropna()) == 0])],
            ['', ''],
            ['Data Quality Thresholds', ''],
            ['Minimum Data Points for Analysis', '2'],
            ['Minimum Years for Pairwise Analysis', '2'],
            ['', ''],
            ['Analysis Types Available', ''],
            ['Original MK Analysis', 'Basic trend analysis for all wells with data'],
            ['Seasonal MK Analysis', 'Individual + pairwise seasonal analysis'],
            ['Duration Analysis', 'Consecutive trend periods'],
            ['Low Flow Analysis', 'Below percentile thresholds'],
            ['Yearly Statistics', 'Annual min/max/mean values'],
            ['Hydrological Year Analysis', 'July-June year statistics']
        ]
        
        for row in overview_data:
            ws_overview.append(row)
        
        # Style overview sheet
        for row in ws_overview.iter_rows(min_row=1, max_row=len(overview_data)):
            for cell in row:
                if cell.value and 'Analysis Summary Report' in str(cell.value):
                    cell.font = Font(bold=True, size=14)
                elif cell.value and any(keyword in str(cell.value) for keyword in ['Total Wells', 'Wells with Data', 'Wells without Data']):
                    cell.font = Font(bold=True)
                    cell.fill = PatternFill(start_color="E6F3FF", end_color="E6F3FF", fill_type="solid")
        
        # 2. DATA QUALITY SHEET
        ws_data = wb.create_sheet("Data_Quality")
        df_data = pd.DataFrame(summary_data)
        
        # Add data quality assessment
        df_data['Data_Quality'] = df_data['Total_Data_Points'].apply(lambda x: 
            'Excellent' if x >= 1000 else
            'Good' if x >= 500 else
            'Fair' if x >= 100 else
            'Poor' if x >= 10 else
            'Very Poor' if x > 0 else 'No Data'
        )
        
        df_data['Analysis_Ready'] = df_data['Total_Data_Points'].apply(lambda x: 'Yes' if x >= 2 else 'No')
        
        # Write to Excel
        for r in dataframe_to_rows(df_data, index=False, header=True):
            ws_data.append(r)
        
        # Style data quality sheet
        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        
        for cell in ws_data[1]:
            cell.font = header_font
            cell.fill = header_fill
        
        # Auto-adjust column widths
        for column in ws_data.columns:
            max_length = 0
            column_letter = column[0].column_letter
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 50)
            ws_data.column_dimensions[column_letter].width = adjusted_width
        
        # 3. ANALYSIS STATUS SHEET (to be populated by each analysis)
        ws_status = wb.create_sheet("Analysis_Status")
        status_headers = [
            'Well', 'Original_Name', 'Original_MK_Status', 'Original_MK_Reason',
            'Seasonal_MK_Status', 'Seasonal_MK_Reason', 'Hydrological_Year_Status', 'Hydrological_Year_Reason',
            'Time_Evolution_Status', 'Time_Evolution_Reason', 'Duration_Low_Flow_Status', 'Duration_Low_Flow_Reason',
            'Yearly_Stats_Status', 'Yearly_Stats_Reason', 'GWL_Plots_Status', 'GWL_Plots_Reason'
        ]
        ws_status.append(status_headers)
        
        # Style status sheet
        for cell in ws_status[1]:
            cell.font = header_font
            cell.fill = header_fill
        
        # Save the workbook
        excel_file = os.path.join(save_directory, "Analysis_Summary_Report.xlsx")
        wb.save(excel_file)
        
        print(f"✓ Created comprehensive analysis summary: {excel_file}")
        return excel_file
        
    except Exception as e:
        print(f"Error creating analysis summary Excel: {e}")
        traceback.print_exc()
        return None


def update_analysis_status_excel(excel_file, analysis_type, well_results):
    """
    Update the analysis status sheet in the Excel file with results from a specific analysis.
    
    Args:
        excel_file: Path to the Excel file
        analysis_type: Type of analysis (e.g., 'Original_MK', 'Seasonal_MK', 'Duration', 'Low_Flow', 'Yearly_Stats')
        well_results: Dictionary with well names as keys and status/reason as values
    """
    try:
        # Load the workbook
        wb = load_workbook(excel_file)
        
        # Get the Analysis_Status sheet
        if 'Analysis_Status' not in wb.sheetnames:
            print(f"⚠ Analysis_Status sheet not found in {excel_file}")
            return
        
        ws = wb['Analysis_Status']
        
        # Find the status and reason columns for this analysis type
        status_col = None
        reason_col = None
        
        for col_idx, cell in enumerate(ws[1], 1):
            if f'{analysis_type}_Status' in str(cell.value):
                status_col = col_idx
            elif f'{analysis_type}_Reason' in str(cell.value):
                reason_col = col_idx
        
        # Handle special cases for analysis type names
        if status_col is None or reason_col is None:
            # Try alternative naming patterns
            for col_idx, cell in enumerate(ws[1], 1):
                if analysis_type == 'Duration_Low_Flow' and 'Duration_Low_Flow_Status' in str(cell.value):
                    status_col = col_idx
                elif analysis_type == 'Duration_Low_Flow' and 'Duration_Low_Flow_Reason' in str(cell.value):
                    reason_col = col_idx
        
        if status_col is None or reason_col is None:
            print(f"⚠ Could not find columns for {analysis_type} in Analysis_Status sheet")
            return
        
        # Update the status for each well
        for row_idx in range(2, ws.max_row + 1):  # Start from row 2 (skip header)
            well_name = ws.cell(row=row_idx, column=1).value  # Well column
            
            if well_name in well_results:
                # Update status
                ws.cell(row=row_idx, column=status_col).value = well_results[well_name]['status']
                ws.cell(row=row_idx, column=reason_col).value = well_results[well_name]['reason']
                
                # Color code the status
                status_cell = ws.cell(row=row_idx, column=status_col)
                if well_results[well_name]['status'] == 'SUCCESS':
                    status_cell.fill = PatternFill(start_color="90EE90", end_color="90EE90", fill_type="solid")  # Light green
                elif well_results[well_name]['status'] == 'SKIPPED':
                    status_cell.fill = PatternFill(start_color="FFB6C1", end_color="FFB6C1", fill_type="solid")  # Light red
                elif well_results[well_name]['status'] == 'ERROR':
                    status_cell.fill = PatternFill(start_color="FFD700", end_color="FFD700", fill_type="solid")  # Gold
        
        # Save the workbook
        wb.save(excel_file)
        print(f"✓ Updated {analysis_type} status in {excel_file}")
        
    except Exception as e:
        print(f"Error updating analysis status Excel: {e}")
        traceback.print_exc()


__all__ = [
    'create_analysis_summary_excel',
    'update_analysis_status_excel'
]
